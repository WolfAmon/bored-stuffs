
import openpyxl

wb = openpyxl.load_workbook("example.xlsx")
print(type(wb))
print(wb.get_sheet_names())
sheet = wb.get_sheet_by_name('Hoja1')

# Accessing the cell values of the workbook
print(sheet['A1'])
print(sheet['A1'].value)

c = sheet['B1']
print(c.value)

print('Row ' + str(c.row) + ', Column ' + str(c.column) + ' is ' + c.value)
print('Cell ' + c.coordinate + ' is ' + c.value)
print(sheet['C1'].value)

print(sheet.cell(row=1, column=2))
print(sheet.cell(row=1, column=2).value)

for i in range(1, 8, 2):
    print(i, sheet.cell(row=i, column=2).value)

sheetNew = wb.get_sheet_by_name('Hoja1')
print(sheetNew.max_row)
print(sheetNew.max_column)

# print(sheet)
# print(type(sheet))

# print(sheet.title)

# anotherSheet = wb.active

# print(anotherSheet)

